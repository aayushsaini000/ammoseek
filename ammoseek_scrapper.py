import cloudscraper
from multiprocessing import Process
import json
import pandas as pd
import openpyxl
from datetime import datetime
import time
from config import payload, cookie
import logging
from logging.handlers import RotatingFileHandler
import sys


class AmmoseekScrapper:
  scraper = cloudscraper.create_scraper()
  mgf_dict = {
    'Ammo Inc': {'id': 998, 'sheet_name': 'ammo_inc'},
    'CCI': {'id': 291, 'sheet_name': 'cci'},
    'Federal': {'id': 5, 'sheet_name': 'federal'},
    'Remington': {'id': 10, 'sheet_name': 'remington'}}

  # mgf_dict = {'Ammo Inc': {'id': 998, 'sheet_name': 'ammo_inc'}}
  
  def getGunTypeAmmo(self, excelfile, gunType, base_url_rifle):
    """
    Request Api for each mfg and gun type
    """
    headers = {
      'Cookie': cookie,
      'Content-Type': 'application/x-www-form-urlencoded'
    }
    for mfg_name, dict_obj  in self.mgf_dict.items():
      url = f"{base_url_rifle}{mfg_name}"
      payload_dict = payload.copy()
      payload_dict['gun'] = gunType
      payload_dict['mfg'] = dict_obj['id']
      start = 0
      flag = True
      draw = 1
      while(flag):
        #logger.info(f"While loop flag: {flag}")
        json_res = self.makeRequest(url, headers, payload_dict, start, draw)
        if bool(json_res):
          #logger.info(f"Data found, totaol records are : {json_res['recordsTotal']}")
          # feed excel
          self.feedExcel(excelfile, dict_obj['sheet_name'], json_res)
          #logger.info('Excel Feeded.')
          flag = True
          draw += 1
          start += 100
          time.sleep(5)
        else:
          #logger.info('While else.')
          flag = False
      time.sleep(5)    

  def makeRequest(self, url, headers, payload, start, draw):
    """
    Calls Api
    """
    payload['start'] = start
    payload['draw'] = draw
    #logger.info(f'Calling Api with url: {url}')
    #logger.info(f'Api params draw: {draw}, start: {start}')
    res = self.scraper.post(url, headers=headers, data=payload)
    if res.status_code == 200:
      try:
        json_res = res.json()
        if 'data' in json_res and json_res['data']:
          return json_res
        else:
          #logger.info(f'Api response with no data key found: {res.text}')
          return {}   
      except json.decoder.JSONDecodeError as ex:
        #logger.error(f'Api json deocde error occured: {res.text}')
        sys.exit()
    else:
      #logger.error(f'Api request error code: {res.status_code}')
      #logger.error(f'Api request error response:  {res.text}')
      sys.exit()
  
  def feedExcel(self, excelfile, sheet, json_data):
    """
    Appends data to existed excel sheet
    """
    dt_str = datetime.now().strftime("%m/%d/%Y, %H:%M:%S")
    data_list = []
    for dict_obj in json_data['data']:
      d = dict()
      d['DateTime'] = dt_str
      d['Retailer'] = dict_obj.get('retailer', '')
      d['Description'] = dict_obj.get('descr', '')
      d['Mfg'] = dict_obj.get('mfg', '')
      d['Caliber'] = dict_obj.get('caliber', '')
      d['Grains'] = dict_obj.get('grains', '')
      d['When'] = dict_obj.get('when', '')
      d['Limits'] = dict_obj.get('purchaselimit', '')
      d['Casing'] = dict_obj.get('casing', '')
      d['New'] = dict_obj.get('dr', '')
      price = dict_obj.get('price', '')
      if price:
        price = float(price[1:])
      d['Price'] = price
      d['Rounds'] = dict_obj.get('count', '')
      cp_raw = dict_obj.get('cp', '')
      if cp_raw:
        #logger.info(f'cp: {cp_raw}')
        if '#' in cp_raw:
          cp = float(cp_raw[:-6])/100
          #logger.info(f'After Conversion cents to dollar: {cp}')
        else:
          cp = float(cp_raw[1:])
      else:
        cp = cp_raw
      d['Dollars'] = cp
      data_list.append(d)
    
    df = pd.DataFrame(data_list)
    wb = openpyxl.load_workbook(excelfile)
    sheet_dict = {ws.title:ws for ws in wb.worksheets}
    sheet_obj = sheet_dict[sheet]
    row_count = sheet_obj.max_row
    
    with pd.ExcelWriter(excelfile, engine="openpyxl", mode='a') as writer:
      writer.book = openpyxl.load_workbook(excelfile)
      writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
      df.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=row_count)

if __name__=='__main__':
  rfh = RotatingFileHandler(
    filename='ammoseek.log', 
    mode='a',
    maxBytes=20*1024*1024,
    backupCount=2,
    encoding=None,
    delay=0
  )

  logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(name)-2s {%(pathname)s:%(lineno)d} %(levelname)-4s %(message)s",
    datefmt="%y-%m-%d %H:%M:%S",
    handlers=[
        rfh
    ]
  )

  #logger = logging.getLogger('main')
  ammoseek_rifle_excel_file = 'Rifles.xlsx'
  ammoseek_handgun_excel_file = 'Handguns.xlsx'
  base_url_rifle = "https://ammoseek.com/rifle-ammo/"
  base_url_handgun = "https://ammoseek.com/handgun-ammo/"
  
  obj = AmmoseekScrapper()
  
  p1 = Process(target = obj.getGunTypeAmmo, args=(ammoseek_rifle_excel_file, 'rifle', base_url_rifle))
  p1.start()
  p2 = Process(target = obj.getGunTypeAmmo, args=(ammoseek_handgun_excel_file, 'handgun', base_url_handgun))
  p2.start()
